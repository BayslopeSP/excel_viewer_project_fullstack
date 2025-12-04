from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.contrib.auth.models import User
from .models import ClientFile, ExcelFile
from .models import Sheet
import os
import openpyxl
from .views import safe_cell_value,save_image_to_disk
from .models import ExcelFile, Sheet, SheetEntry, Image
from rest_framework.permissions import AllowAny

from .models import ClientFile, ExcelFile, Sheet, SheetEntry, Image

# import pdfplumber
# import re\\


REQUIRED_SHEETS = [
    "COVER PAGE",
    "TABLE OF CONTENTS",
    "OBJECTIVE & METHODOLOGY",
    "OVERVIEW",
    "PRIMARY RESULTS - MATRIX",
    "PRIMARY RESULTS - MAPPING",
    "PRIMARY RESULTS - BIBLIO",
    "SECONDARY RESULTS",
    "SEARCH STRINGS",
    "DISCLAIMER"
]

def validate_required_sheets(sheetnames):
    sheetnames_upper = [name.strip().upper() for name in sheetnames]
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet.upper() not in sheetnames_upper]
    return missing

def is_sheet_meaningful(ws):
    for row in ws.iter_rows():
        row_data = [cell.value for cell in row if cell.value not in [None, '', ' ']]
        if row_data:
            return True
    return False


import pdfplumber

def extract_pdf_full(pdf_path):
    extracted = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract all tables from this page
            tables = page.extract_tables()
            # Extract all text lines from this page
            text = page.extract_text()
            lines = text.split('\n') if text else []

            # Merge tables and text in order (line by line)
            # pdfplumber does not give table position, so we add all text first, then all tables
            for line in lines:
                extracted.append(line)
            for table in tables:
                extracted.append({"table": table})
    return extracted

class PdfFullExtractView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, file_id):
        try:
            excel_file = ExcelFile.objects.get(id=file_id)
        except ExcelFile.DoesNotExist:
            return Response({"error": "PDF not found"}, status=404)

        pdf_path = excel_file.file.path
        full_data = extract_pdf_full(pdf_path)
        return Response(full_data)

import pdfplumber
import re

def extract_pdf_tabs(pdf_path):
    tabs = []
    current_tab = None
    intro_content = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            print("Page tables:", tables)
            text = page.extract_text()
            if not text:
                continue
            lines = text.split('\n')
            
            table_idx = 0
            for line in lines:
                # Heading detection
                if (
                    re.match(r'^\d+\.\d+ ', line.strip()) or
                    'disclaimer' in line.strip().lower() or
                    'central references' in line.strip().lower() or
                    'feature matrix' in line.strip().lower()
                ):
                    heading = line.strip()
                    if (
                        heading.lower().startswith('intro') or
                        'feature matrix' in heading.lower() or
                        'central references' in heading.lower() or
                        'disclaimer' in heading.lower()
                    ):
                        if current_tab is None and intro_content:
                            tabs.append({"heading": "INTRO", "content": intro_content})
                        current_tab = {"heading": heading, "content": []}
                        tabs.append(current_tab)
                        continue

                # Central References: Result detection
                if current_tab and "central references" in current_tab["heading"].lower():
                    if re.match(r'^\s*result\s+\d+', line.strip(), re.IGNORECASE):
                        current_tab["content"].append({"result_heading": line.strip(), "content": []})
                        continue
                    if current_tab["content"] and table_idx < len(tables):
                        last = current_tab["content"][-1]
                        if isinstance(last, dict) and "content" in last and not last.get("table_added"):
                            last["content"].append({"table": tables[table_idx]})
                            last["table_added"] = True
                            table_idx += 1
                            continue
                    if current_tab["content"]:
                        last = current_tab["content"][-1]
                        if isinstance(last, dict) and "content" in last:
                            last["content"].append(line)
                        else:
                            current_tab["content"].append(line)
                    else:
                        current_tab["content"].append(line)
                    continue

                # Feature Matrix: Table detection
                if current_tab and "feature matrix" in current_tab["heading"].lower():
                    if table_idx < len(tables):
                        current_tab["content"].append({"table": tables[table_idx]})
                        table_idx += 1
                        continue

                # Normal line append
                if current_tab:
                    current_tab["content"].append(line)
                else:
                    intro_content.append(line)

            # YAHAN TABLES KO ADD KARO (for table in tables ...)
            for table in tables:
                if current_tab and "feature matrix" in current_tab["heading"].lower():
                    current_tab["content"].append({"table": table})
                elif current_tab and "central references" in current_tab["heading"].lower():
                    current_tab["content"].append({"table": table})
                elif current_tab:
                    current_tab["content"].append({"table": table})
                elif intro_content is not None:
                    intro_content.append({"table": table})

    if intro_content and (not tabs or tabs[0]["heading"] != "INTRO"):
        tabs.insert(0, {"heading": "INTRO", "content": intro_content})

    return tabs


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from .models import ExcelFile

class PdfTabsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, file_id):
        try:
            excel_file = ExcelFile.objects.get(id=file_id)
        except ExcelFile.DoesNotExist:
            return Response({"error": "PDF not found"}, status=404)

        pdf_path = excel_file.file.path
        tabs = extract_pdf_tabs(pdf_path)
        for tab in tabs:
            if "feature matrix" in tab["heading"].lower():
                print("Feature Matrix Tab Content:", tab["content"])
        return Response(excel_file.pdf_full or [])



# class AdminFileUploadView(APIView):
#     permission_classes = [permissions.IsAdminUser]

#     def post(self, request):
#         user_id = request.data.get('user_id')
#         file_obj = request.FILES.get('file')
#         file_name = file_obj.name if file_obj else None

#         if not user_id or not file_obj:
#             return Response({'error': 'user_id and file are required.'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             user = User.objects.get(id=user_id)
#         except User.DoesNotExist:
#             return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)

#         filename = file_obj.name
#         ext = os.path.splitext(filename)[1].lower()

#         if ext in ['.xlsx', '.xls']:
#             try:
#                 wb = openpyxl.load_workbook(file_obj, data_only=False)
#             except Exception as e:
#                 return Response({"error": f"Invalid Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

#             # Reset file pointer for saving
#             file_obj.seek(0)
#             excel_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
#             client_file = ClientFile.objects.create(
#                 user=user,
#                 file=file_obj,
#                 file_name=filename,
#                 excel_file=excel_file
#             )

#             response_data = {"type": "excel", "excel_file_id": excel_file.id, "client_file_id": client_file.id, "sheets": []}

#             for sheet_name in wb.sheetnames:
#                 ws = wb[sheet_name]
#                 sheet = Sheet.objects.create(
#                     name=sheet_name,
#                     excel_file=excel_file,
#                     merged_cells=[str(merge) for merge in ws.merged_cells.ranges],
#                     column_widths={col_letter: ws.column_dimensions[col_letter].width for col_letter in ws.column_dimensions},
#                     row_heights={row_num: ws.row_dimensions[row_num].height for row_num in ws.row_dimensions},
#                 )

#                 rows = []
#                 for row in ws.iter_rows():
#                     row_data = [safe_cell_value(cell, sheet_name) for cell in row]
#                     rows.append(row_data)

#                 SheetEntry.objects.create(sheet=sheet, row_data=rows)

#                 # Images
#                 image_instances = []
#                 images_json = []
#                 for img in getattr(ws, '_images', []):
#                     image_instance = save_image_to_disk(img)
#                     if image_instance:
#                         anchor = img.anchor._from
#                         row = anchor.row
#                         col = anchor.col
#                         try:
#                             cell = ws.cell(row=row + 1, column=col + 1)
#                             cell_value = str(cell.value).strip() if cell.value else ''
#                         except Exception as e:
#                             cell_value = ''
#                         image_instance.row = row
#                         image_instance.column = col
#                         image_instance.save()
#                         image_instances.append(image_instance)
#                         images_json.append({
#                             "url": request.build_absolute_uri(f"/media/{image_instance.image.name}"),
#                             "row": row + 1,
#                             "column": col + 1,
#                             "cell_value": cell_value,
#                         })

#                 sheet.images.set(image_instances)
#                 sheet.save()

#                 response_data["sheets"].append({
#                     "id": sheet.id,
#                     "name": sheet.name,
#                     "columns": rows[0] if rows else [],
#                     "rows": rows[1:] if rows else [],
#                     "merged_cells": sheet.merged_cells,
#                     "column_widths": sheet.column_widths,
#                     "row_heights": sheet.row_heights,
#                     "images": images_json,
#                 })

#             return Response(response_data, status=status.HTTP_201_CREATED)

#         elif ext == '.pdf':
#             # --- PDF Handling ---
#             from PyPDF2 import PdfReader
#             file_obj.seek(0)
#             pdf_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
#             client_file = ClientFile.objects.create(
#                 user=user,
#                 file=file_obj,
#                 file_name=filename,
#                 excel_file=pdf_file
#             )
#             try:
#                 reader = PdfReader(file_obj)
#                 num_pages = len(reader.pages)
#                 pdf_info = reader.metadata
#             except Exception as e:
#                 num_pages = None
#                 pdf_info = {}

#             pdf_path = pdf_file.file.path
#             full_data = extract_pdf_full(pdf_path)  # <-- Yeh line
#             # tabs = extract_pdf_tabs(pdf_path)
#             # pdf_file.pdf_tabs = tabs
#             pdf_file.pdf_full = full_data
#             pdf_file.save()

#             file_url = request.build_absolute_uri(pdf_file.file.url) if pdf_file.file else None
#             return Response({
#                 "type": "pdf",
#                 "pdf_file_id": pdf_file.id,
#                 "client_file_id": client_file.id,
#                 "file_name": pdf_file.file_name,
#                 "file_url": pdf_file.file.url if pdf_file.file else "",
#                 "num_pages": num_pages,
#                 "pdf_info": {k: str(v) for k, v in pdf_info.items()} if pdf_info else {},
#                 # "pdf_tabs": tabs,
#                 "pdf_full": full_data,
#                 "message": "PDF uploaded successfully."
#             }, status=status.HTTP_201_CREATED)

#         else:
#             return Response({"error": "Unsupported file type."}, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.contrib.auth.models import User
from .models import ClientFile, ExcelFile, Sheet, SheetEntry
from .views import safe_cell_value, save_image_to_disk
import os
import openpyxl
from PyPDF2 import PdfReader

class AdminFileUploadView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request):
        user_id = request.data.get('user_id')
        file_obj = request.FILES.get('file')
        file_name = file_obj.name if file_obj else None

        if not user_id or not file_obj:
            return Response({'error': 'user_id and file are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)

        filename = file_obj.name
        ext = os.path.splitext(filename)[1].lower()

        if ext in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(file_obj, data_only=False)
            except Exception as e:
                return Response({"error": f"Invalid Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

            # ➤ Validate required sheets
            missing_sheets = validate_required_sheets(wb.sheetnames)
            if missing_sheets:
                return Response({
                    "error": "Excel is missing required sheets.",
                    "missing_sheets": missing_sheets
                }, status=status.HTTP_400_BAD_REQUEST)

            # ➤ Validate file has meaningful data
            valid_sheet_found = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if is_sheet_meaningful(ws):
                    valid_sheet_found = True
                    break

            if not valid_sheet_found:
                return Response({"error": "Excel file does not contain any meaningful data."}, status=status.HTTP_400_BAD_REQUEST)

            # Reset file pointer for saving
            file_obj.seek(0)
            excel_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
            client_file = ClientFile.objects.create(
                user=user,
                file=file_obj,
                file_name=filename,
                excel_file=excel_file
            )

            response_data = {
                "type": "excel",
                "excel_file_id": excel_file.id,
                "client_file_id": client_file.id,
                "sheets": []
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                sheet = Sheet.objects.create(
                    name=sheet_name,
                    excel_file=excel_file,
                    merged_cells=[str(merge) for merge in ws.merged_cells.ranges],
                    column_widths={col_letter: ws.column_dimensions[col_letter].width for col_letter in ws.column_dimensions},
                    row_heights={row_num: ws.row_dimensions[row_num].height for row_num in ws.row_dimensions},
                )

                rows = []
                for row in ws.iter_rows():
                    row_data = [safe_cell_value(cell, sheet_name) for cell in row]
                    rows.append(row_data)

                SheetEntry.objects.create(sheet=sheet, row_data=rows)

                # ➤ Handle Images
                image_instances = []
                images_json = []
                for img in getattr(ws, '_images', []):
                    image_instance = save_image_to_disk(img)
                    if image_instance:
                        anchor = img.anchor._from
                        row = anchor.row
                        col = anchor.col
                        try:
                            cell = ws.cell(row=row + 1, column=col + 1)
                            cell_value = str(cell.value).strip() if cell.value else ''
                        except:
                            cell_value = ''
                        image_instance.row = row
                        image_instance.column = col
                        image_instance.save()
                        image_instances.append(image_instance)
                        images_json.append({
                            "url": request.build_absolute_uri(f"/media/{image_instance.image.name}"),
                            "row": row + 1,
                            "column": col + 1,
                            "cell_value": cell_value,
                        })

                sheet.images.set(image_instances)
                sheet.save()

                response_data["sheets"].append({
                    "id": sheet.id,
                    "name": sheet.name,
                    "columns": rows[0] if rows else [],
                    "rows": rows[1:] if rows else [],
                    "merged_cells": sheet.merged_cells,
                    "column_widths": sheet.column_widths,
                    "row_heights": sheet.row_heights,
                    "images": images_json,
                })

            return Response(response_data, status=status.HTTP_201_CREATED)

        elif ext == '.pdf':
            # --- PDF Handling ---
            file_obj.seek(0)
            pdf_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
            client_file = ClientFile.objects.create(
                user=user,
                file=file_obj,
                file_name=filename,
                excel_file=pdf_file
            )

            try:
                reader = PdfReader(file_obj)
                num_pages = len(reader.pages)
                pdf_info = reader.metadata
            except Exception as e:
                num_pages = None
                pdf_info = {}

            pdf_path = pdf_file.file.path
            full_data = extract_pdf_full(pdf_path)
            pdf_file.pdf_full = full_data
            pdf_file.save()

            file_url = request.build_absolute_uri(pdf_file.file.url) if pdf_file.file else None
            return Response({
                "type": "pdf",
                "pdf_file_id": pdf_file.id,
                "client_file_id": client_file.id,
                "file_name": pdf_file.file_name,
                "file_url": file_url,
                "num_pages": num_pages,
                "pdf_info": {k: str(v) for k, v in pdf_info.items()} if pdf_info else {},
                "pdf_full": full_data,
                "message": "PDF uploaded successfully."
            }, status=status.HTTP_201_CREATED)

        else:
            return Response({"error": "Unsupported file type."}, status=status.HTTP_400_BAD_REQUEST)



from rest_framework import generics, permissions
from .models import ClientFile
from .serializers import ClientFileSerializer

class ClientFileListView(generics.ListAPIView):
    serializer_class = ClientFileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return ClientFile.objects.filter(user=self.request.user)






# from rest_framework import generics
from .serializers import RegisterSerializer
from django.contrib.auth.models import User
from rest_framework import generics, permissions

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]



from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['is_staff'] = user.is_staff
        return token

    def validate(self, attrs):
        data = super().validate(attrs)
        data['is_staff'] = self.user.is_staff
        return data

class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from django.contrib.auth.models import User

class ClientListView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        users = User.objects.all()
        data = [
            {
                "id": user.id,
                "username": user.username,
                "email": user.email,
            }
            for user in users
        ]
        return Response(data)



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from .models import ClientFile

class AdminClientFilesView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, user_id):
        files = ClientFile.objects.filter(user_id=user_id)
        data = [
            {
                "id": f.id,
                "file_name": f.file_name,
                "file": request.build_absolute_uri(f.file.url),
                "uploaded_at": f.uploaded_at
            }
            for f in files
        ]
        return Response(data)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from .models import ExcelFile, Sheet, SheetEntry
import time, logging
from collections import defaultdict
from django.views.decorators.cache import cache_page
from django.utils.decorators import method_decorator

logger = logging.getLogger(__name__)

@method_decorator(cache_page(60 * 2), name='dispatch')  # 2-min cache
class ClientFileSheetsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        start = time.time()

        user = request.user
        all_data = []

        user_files = ExcelFile.objects.filter(clientfile__user=user).only('id', 'file_name', 'file')

        base_url = request.build_absolute_uri('/')

        sheets = Sheet.objects.filter(excel_file__in=user_files).prefetch_related('images')
        sheet_map = defaultdict(list)
        for sheet in sheets:
            sheet_map[sheet.excel_file_id].append(sheet)

        entries = SheetEntry.objects.filter(sheet__in=sheets).order_by('sheet_id', '-date_mapped')
        entry_map = {}
        for entry in entries:
            if entry.sheet_id not in entry_map:
                entry_map[entry.sheet_id] = entry

        for excel_file in user_files:
            file_data = {
                "id": excel_file.id,
                "file_name": excel_file.file_name,
                "file_url": base_url + excel_file.file.name if excel_file.file else "",
                "sheets": [],
            }

            sheets = sheet_map.get(excel_file.id, [])

            for sheet in sheets:
                entry = entry_map.get(sheet.id)

                sheet_data = {
                    "id": sheet.id,
                    "name": sheet.name,
                    "columns": [],
                    "rows": [],
                    "images": [],
                }

                if entry and entry.row_data:
                    rows = entry.row_data
                    sheet_data["columns"] = rows[0] if rows else []
                    sheet_data["rows"] = rows[1:4] if len(rows) > 1 else []

                sheet_data["images"] = [
                    {
                        "url": base_url + img.image.name,
                        "row": img.row + 1 if img.row is not None else None,
                        "column": img.column + 1 if img.column is not None else None
                    }
                    for img in sheet.images.all()
                    if img.image and img.image.name
                ]

                file_data["sheets"].append(sheet_data)

            all_data.append(file_data)

        total_time = time.time() - start
        logger.info(f"[SYNC] `/client/files/` took {total_time:.2f}s for user {user.username}")
        return Response(all_data)






from django.http import FileResponse, Http404
import mimetypes
import os
from django.conf import settings

def serve_media(request, path):
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if not os.path.exists(file_path):
        raise Http404("File not found")
    mime_type, _ = mimetypes.guess_type(file_path)
    return FileResponse(open(file_path, "rb"), content_type=mime_type or "application/octet-stream")



# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status, permissions
from .models import ClientFile, ExcelFile

class DeleteClientFileView(APIView):
    permission_classes = [permissions.IsAdminUser]  # Only admin can delete

    def delete(self, request, file_id):
        try:
            client_file = ClientFile.objects.get(id=file_id)
        except ClientFile.DoesNotExist:
            return Response({"error": "File not found."}, status=status.HTTP_404_NOT_FOUND)

        # Optionally delete related ExcelFile as well
        if client_file.excel_file:
            client_file.excel_file.delete()

        # Delete the main file
        if client_file.file and client_file.file.path and os.path.exists(client_file.file.path):
            os.remove(client_file.file.path)

        client_file.delete()

        return Response({"message": "File deleted successfully."}, status=status.HTTP_204_NO_CONTENT)