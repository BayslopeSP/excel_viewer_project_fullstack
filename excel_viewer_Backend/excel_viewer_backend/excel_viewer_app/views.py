import datetime
import openpyxl
import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from django.core.files import File
from rest_framework import status
from .models import ExcelFile, Sheet, SheetEntry, Image
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.core.files.temp import NamedTemporaryFile
from .utils import parse_sheet_data, filter_mapping_rows_by_claim
from io import BytesIO
import base64
import os
from django.conf import settings
import uuid
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# Helper function to save image to disk
def save_image_to_disk(img: OpenpyxlImage):
    try:
        # Create a directory for images if it doesn't exist
        image_dir = os.path.join(settings.MEDIA_ROOT, 'images')
        os.makedirs(image_dir, exist_ok=True)

        # Generate a unique filename for the image using UUID
        image_filename = f"{uuid.uuid4()}.png"
        image_path = os.path.join(image_dir, image_filename)

        # Get image data
        image_stream = img._data()  # This returns a bytes object

        if not image_stream:
            print("No image data found.")
            return None

        # Open the image using PIL
        pil_image = PILImage.open(BytesIO(image_stream))

        # Save image to disk
        pil_image.save(image_path, format="PNG")

        # Save to DB
        image_instance = Image.objects.create(
            image=File(open(image_path, 'rb'), name=image_filename)
        )

        return image_instance
    except Exception as e:
        print(f"Error saving image: {e}")
        return None


def safe_cell_value(cell, sheet_name):
    value = cell.value
    if isinstance(value, (datetime.datetime, datetime.date)):
        value = value.strftime('%Y-%m-%d')

    hyperlink = cell.hyperlink.target if cell.hyperlink else None
    target_sheet = None
    target_column = None

    # If there's a hyperlink, extract target sheet and column info (assuming format like 'Sheet2!A1')
    if hyperlink:
        link_parts = hyperlink.split("!")
        if len(link_parts) == 2:
            target_sheet = link_parts[0]
            target_column = link_parts[1]

    return {
        "value": value,
        "font_color": cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
        "fill_color": cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor.type == "rgb" else None,
        "hyperlink": hyperlink,
        "target_sheet": target_sheet,
        "target_column": target_column,
        "bold": cell.font.bold,
        "italic": cell.font.italic,
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical
        },
        "is_merged": cell.coordinate in cell.parent.merged_cells,
        "checkbox": str(value).strip() in ['✓', '☑']
    }



import os
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from PyPDF2 import PdfReader  # pip install PyPDF2

from .models import ExcelFile, Sheet, SheetEntry  # adjust import as per your project
# from .utils import safe_cell_value, save_image_to_disk  # adjust import as per your project

class ExcelUploadView(APIView):
    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name
        ext = os.path.splitext(filename)[1].lower()

        if ext in ['.xlsx', '.xls']:
            # --- Excel Handling ---
            try:
                wb = openpyxl.load_workbook(file_obj, data_only=False)
            except Exception as e:
                return Response({"error": f"Invalid Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

            excel_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
            response_data = {"excel_file_id": excel_file.id, "sheets": []}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet = Sheet.objects.create(
                    name=sheet_name, 
                    excel_file=excel_file,
                    merged_cells=[str(merge) for merge in ws.merged_cells.ranges],
                    column_widths={ 
                        col_letter: ws.column_dimensions[col_letter].width 
                        for col_letter in ws.column_dimensions
                    },
                    row_heights={ 
                        row_num: ws.row_dimensions[row_num].height 
                        for row_num in ws.row_dimensions
                    },
                )

                rows = []
                for row in ws.iter_rows():
                    row_data = [safe_cell_value(cell, sheet_name) for cell in row]
                    rows.append(row_data)

                # Creating SheetEntry for storing row data
                SheetEntry.objects.create(sheet=sheet, row_data=rows)

                # Process images from the Excel sheet and save them
                image_instances = []  # model instances for DB save
                images_json = []
                for img in getattr(ws, '_images', []):
                    image_instance = save_image_to_disk(img)
                    if image_instance:
                        anchor = img.anchor._from  # only works for openpyxl images
                        row = anchor.row       # 0-based
                        col = anchor.col       # 0-based

                        try:
                            cell = ws.cell(row=row + 1, column=col + 1)  # openpyxl is 1-based
                            cell_value = str(cell.value).strip() if cell.value else ''
                        except Exception as e:
                            print(f"⚠️ Failed reading cell value at row={row+1} col={col+1}: {e}")
                            cell_value = ''

                        image_instance.row = row
                        image_instance.column = col
                        image_instance.save()

                        image_instances.append(image_instance)

                        images_json.append(
                            {
                                "url": image_instance.image.url,
                                "row": row + 1,  # 1-based for frontend
                                "column": col + 1,
                                "cell_value": cell_value,
                            }
                        )

                sheet.images.set(image_instances)
                sheet.save()

                response_data["sheets"].append(
                    {
                        "id": sheet.id,
                        "name": sheet.name,
                        "columns": rows[0] if rows else [],
                        "rows": rows[1:] if rows else [],
                        "merged_cells": sheet.merged_cells,
                        "column_widths": sheet.column_widths,
                        "row_heights": sheet.row_heights,
                        "images": images_json,
                    }
                )

            return Response(response_data, status=status.HTTP_200_OK)

        elif ext == '.pdf':
            # --- PDF Handling ---
            pdf_file = ExcelFile.objects.create(file_name=filename, file=file_obj)
            # Optionally extract PDF metadata
            try:
                file_obj.seek(0)
                reader = PdfReader(file_obj)
                num_pages = len(reader.pages)
                pdf_info = reader.metadata
            except Exception as e:
                num_pages = None
                pdf_info = {}

            return Response({
                "pdf_file_id": pdf_file.id,
                "file_name": pdf_file.file_name,
                "num_pages": num_pages,
                "pdf_info": {k: str(v) for k, v in pdf_info.items()} if pdf_info else {},
                "message": "PDF uploaded successfully."
            }, status=status.HTTP_200_OK)

        else:
            return Response({"error": "Unsupported file type."}, status=status.HTTP_400_BAD_REQUEST)


# GET all Excel files and their sheets
class SheetListView(APIView):
    def get(self, request):
        all_data = []

        # Sare ExcelFile fetch karo (Excel + PDF dono)
        all_files = ExcelFile.objects.all()

        for excel_file in all_files:
            # Sare sheets fetch karo
            sheets = Sheet.objects.filter(excel_file=excel_file)
            file_data = {
                "id": excel_file.id,
                "file_name": excel_file.file_name,
                "sheets": [],
                "file_url": excel_file.file.url if excel_file.file else "",
            }

            for sheet in sheets:
                entries = SheetEntry.objects.filter(sheet=sheet).order_by("-date_mapped")
                images = sheet.images.all()
                images_list = [
                    {
                        "url": image.image.url,
                        "row": image.row + 1,
                        "column": image.column + 1,
                    }
                    for image in images
                    if image.row is not None and image.column is not None
                ]

                sheet_data = {
                    "id": sheet.id,
                    "name": sheet.name,
                    "columns": [],
                    "rows": [],
                    "images": images_list,
                }

                if entries.exists():
                    for entry in entries:
                        rows = entry.row_data or []
                        sheet_data["columns"] = rows[0] if rows else []
                        sheet_data["rows"] = rows[1:] if rows else []

                file_data["sheets"].append(sheet_data)

            all_data.append(file_data)

        return Response(all_data)



# Add new view to handle checkbox click
class UpdateMappingDateView(APIView):
    def post(self, request):
        sheet_id = request.data.get('sheet_id')
        row_index = request.data.get('row_index')

        if not sheet_id or not row_index:
            return Response({"error": "sheet_id and row_index are required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            sheet = Sheet.objects.get(id=sheet_id)
            entry = SheetEntry.objects.filter(sheet=sheet).first()

            if entry:
                # Update the date_mapped field
                entry.date_mapped = timezone.now()
                entry.save()

                return Response({"message": "Date updated successfully"}, status=status.HTTP_200_OK)
            else:
                return Response({"error": "No entry found for this sheet"}, status=status.HTTP_404_NOT_FOUND)

        except Sheet.DoesNotExist:
            return Response({"error": "Sheet not found"}, status=status.HTTP_404_NOT_FOUND)


class SpecificSheetView(APIView):
    def get(self, request, file_id):
        try:
            excel_file = ExcelFile.objects.get(id=file_id)
        except ExcelFile.DoesNotExist:
            return Response({"error": "Excel file not found"}, status=status.HTTP_404_NOT_FOUND)

        file_data = {
            "id": excel_file.id,
            "file_name": excel_file.file_name,
            "file_url": request.build_absolute_uri(excel_file.file.url) if excel_file.file else "",
            "sheets": [],
        }

        sheets = Sheet.objects.filter(excel_file=excel_file)
        for sheet in sheets:
            entries = SheetEntry.objects.filter(sheet=sheet)
            images_list = [
                {
                    "url": image.image.url,
                    "row": image.row + 1,
                    "column": image.column + 1
                }
                for image in sheet.images.all()
                if image.row is not None and image.column is not None
            ]
            sheet_data = {
                "id": sheet.id,
                "name": sheet.name,
                "columns": [],
                "rows": [],
                "images": images_list,
                "merged_cells": sheet.merged_cells,
                "column_widths": sheet.column_widths,
                "row_heights": sheet.row_heights,
            }
            if entries.exists():
                for entry in entries:
                    rows = entry.row_data or []
                    if rows:
                        sheet_data["columns"] = rows[0]
                        sheet_data["rows"] = rows[1:]
            file_data["sheets"].append(sheet_data)

        return Response(file_data, status=status.HTTP_200_OK)

# DELETE Excel file and all linked sheets/data
class DeleteExcelFileView(APIView):
    def delete(self, request, file_id):
        try:
            excel_file = ExcelFile.objects.get(id=file_id)
            excel_file.delete()
            return Response({"message": "Excel file and all related data deleted successfully."}, status=status.HTTP_200_OK)
        except ExcelFile.DoesNotExist:
            return Response({"error": "Excel file not found."}, status=status.HTTP_404_NOT_FOUND)

# Update the sheet (if necessary)
class UpdateSheetView(APIView):
    def put(self, request, sheet_id):
        try:
            sheet = Sheet.objects.get(id=sheet_id)
            sheet.name = request.data.get("name", sheet.name)
            sheet.save()
            return Response({"message": "Sheet updated successfully."}, status=status.HTTP_200_OK)
        except Sheet.DoesNotExist:
            return Response({"error": "Sheet not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# Add new view to filter mapping rows by claim
class FilterMappingView(APIView):
    def get(self, request, file_id, claim_id):
        try:
            excel_file = ExcelFile.objects.get(id=file_id)
            sheet = Sheet.objects.filter(excel_file=excel_file, name='Mapping').first()
            
            if not sheet:
                return Response({"error": "Mapping sheet not found"}, status=status.HTTP_404_NOT_FOUND)
                
            entries = SheetEntry.objects.filter(sheet=sheet)
            sheet_data = {
                "columns": entries.first().row_data[0] if entries.exists() else [],
                "rows": [row for entry in entries for row in entry.row_data[1:]]
            }
            
            filtered_rows = filter_mapping_rows_by_claim(sheet_data, claim_id)
            return Response({"rows": filtered_rows})
            
        except ExcelFile.DoesNotExist:
            return Response({"error": "Excel file not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




# ----------------
